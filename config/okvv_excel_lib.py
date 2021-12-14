import pandas as pd
import os


from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Alignment, Side
import re
def add_new_orders(new_row,ws):
    '''
    add_new_orderns(new_row=list, ws=openpyxl.worksheet)

    appends a new row, stored as a ppython list, to a openpyxl worksheet instance
    '''
    order_df=pd.DataFrame([new_row])
    '''
    order_df ist ein Dataframe welches pro Zeile die Daten einer neuen Bestellung enthält.
    Die Zeile hat folgende Struktur:
    (BestNr., Nachname, Vorname, email, Tel., KartenNoSa, KartenErSa, KartenNoSo, KartenErSo, Bemerkung, Datum)
    '''
    order_df['erste']=[1 for x in range(len(order_df))]
    order_df['geänderte']=[None for x in range(len(order_df))]
    order_df['alte']=[None for x in range(len(order_df))]
    columns=list(order_df.columns)
    columns=columns[:10]+columns[-3:]+[columns[10]]
    order_df=order_df[columns]
    
    border=Side(border_style="thin", color="000000")
    alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=False,
                     shrink_to_fit=False,
                     indent=0)
    for r in dataframe_to_rows(order_df, index=False, header=False):
        row_nr=int(re.findall('[1-9][0-9]*',r[0])[0])+3
        ws.append(r)
        #Formatting Vertical lines
        for col in list('EIJMN'):
            cell=ws[col+str(row_nr)]
            cell.border=Border(right=border)
        #Formating Alignment for BestNr Column
        a=ws['A'+str(row_nr)]
        a.alignment=Alignment(horizontal='right')
        
    #updating Ticketsum formula in cells
    for col, nr in list(zip(['F','G','H','I'],[6,7,8,9])):
        ws.cell(column=nr,row=3,value='=SUM({0}4:{0}999)-SUMIF($M$4:$M$130,1,{0}4:{0}999)'.format(col))
        
    #updating Version-sum Cells
    for col, nr in list(zip(['K','L','M'],[11,12,13])):
        cell=ws[col+'3']
        ws.cell(column=nr,row=3,value='=SUM({0}4:{0}999)'.format(col))



import datetime
def data_extractor(fname='1.csv'):
    '''
    data_extractor(fname)
    fname ist ein der Dateiname des csv Formulars im Formulare ORdner.
    Als Input wird der Output der pd.read_csv() funktion erwartet mit der Bestellungs CSV als Input.
    '''
    data={}
    df=pd.read_csv('./Formulare/{}'.format(fname), squeeze = True)
    df=df.str.extract('~(.*)')[0][:10]
    data['anrede']=df[0].capitalize()
    data['nachname']=df[1].capitalize()
    data['vorname']=df[2].capitalize()
    data['email']=df[3]
    data['tel']=df[4]
    for i in range(len(df)): #Falls keine Karten für einen Tag bestellt wurden soll der Eintrag mit 0 gelistet werden
        if df[i]=='':
            df[i]=0
    data['sa_norm']=int(df[5])
    data['sa_erm']=int(df[6])
    data['so_norm']=int(df[7])
    data['so_erm']=int(df[8])
    if df[9]=='M�chten Sie uns noch etwas mitteilen?':
        data['Bemerkung']=''
    else:
        data['Bemerkung']=df[9]
    data['datum']=datetime.datetime.today().strftime("%d.%m.%Y")
    return data

def get_os_dir_sep():
    osIndicator=-1 #0=Win, 1=OSX | Dateipfade werden auf OSX und Windows unterschiedlich geschrieben
    osDirSeperator=''
    try:
        os.listdir('.\\Formulare')
        osIndicator = 0
        osDirSeperator='\\'
        #print('Programm läuft auf Windows')
        return osDirSeperator
    except: 
        try:
            os.listdir('./Formulare')
            osIndicator=1
            osDirSeperator='/'
            #print('Programm läuft auf Linux basiertem Betriebssystem')
            return osDirSeperator
        except: 
            raise Exception('Formulare Ordner konnte nicht gefunden werden. Bitte erstellen Sie im Ordner dieses Skriptes einen Formulare Ordner')



from openpyxl import load_workbook
import regex as re
from shutil import copyfile
def load_excel():
    '''
    This function loads or creates a Excelfile to store the orders in
    Returns a tuple (openpyxl.Workbook, name of opened file)
    '''
    if __name__!='__main__':
        try: #trying to load existing Excel file
            liste=os.listdir() #get all files within Programms Root directory
            r=re.compile('.*xlsx', re.IGNORECASE) # 
            xlnames=list(filter(r.match, liste))
            if len(xlnames)>1:
                print('''Stellen Sie bitte sicher, dass sich im Hauptordner exakt eine Excel Datei befindet
                \n die folgenden Excel-Dateien wurden gefungen
                ''')
            elif len(xlnames)==0:
                print('keine Excel gefunden')

            fname=xlnames[0]

        except: #copying the template from the conf directory in case no excel file was found previously
            print('Es konnte keine Excelliste im Hauptordner des Programmes gefunden werden.')
            fname='okvv_'+input('Bitte geben Sie das aktuelle Semester an (z.B. WS21, SS22 usw.): ') + '.xlsx'
            src='./config/KVV-offentlich-template.xlsx'
            dest=fname
            copyfile(src, dest)
            
        wb=load_workbook(fname)
        print('Die Datei {} wurde geladen'.format(fname))

        return wb, fname
    else: print('Diese Methode muss vom Haupskript aus aufgerufen werden')



